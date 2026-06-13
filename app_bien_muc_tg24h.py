import os
import sys
import json
import logging
import threading
import datetime
import traceback
import re
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import openpyxl
from openpyxl.styles import Font, Alignment
from striprtf.striprtf import rtf_to_text as _rtf_to_text_raw
import google.generativeai as genai
import typing_extensions as typing
from openai import OpenAI

def rtf_to_text(rtf_str):
    """Wrapper: fix ky tu Vietnamese d/D bi sai do RTF CP1252 encoding.
    Thay the ky tu eth (U+00F0) thanh d-stroke (U+0111) sau khi convert."""
    result = _rtf_to_text_raw(rtf_str)
    result = result.replace('\u00f0', 'đ')  # ð → đ
    result = result.replace('\u00d0', 'Đ')  # Ð → Đ
    return result

def apply_tnr_font(ws):
    tnr_font = Font(name='Times New Roman', size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = tnr_font
    
    # Tăng chiều rộng cột B và C lên gấp 3 lần độ rộng mặc định (~8.43 * 3 ≈ 25.3, ta đặt 30 để thoải mái hơn)
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30


def get_green_cf_tag(rtf_raw: str) -> str:
    r"""Tìm thẻ colortbl định nghĩa bảng màu trong file RTF và trả về thẻ \cf tương ứng với màu xanh lá cây."""
    match = re.search(r'\{\\colortbl([^}]+)\}', rtf_raw)
    if not match:
        return r'\cf2'
    colortbl_content = match.group(1)
    colors = colortbl_content.split(';')
    for idx, color_def in enumerate(colors):
        if idx == 0:
            continue
        if r'\green128' in color_def:
            return f'\\cf{idx}'
    return r'\cf2'


def get_color_tags(rtf_raw: str):
    """Parse colortbl để tìm cf tags cho green, black, và red."""
    match = re.search(r'\{\\colortbl([^}]+)\}', rtf_raw)
    green_cf, black_cf, red_cf = None, None, None
    if not match:
        return r'\cf2', r'\cf1', None
    colortbl_content = match.group(1)
    colors = colortbl_content.split(';')
    for idx, color_def in enumerate(colors):
        if idx == 0:
            continue
        r_match = re.search(r'\\red(\d+)', color_def)
        g_match = re.search(r'\\green(\d+)', color_def)
        b_match = re.search(r'\\blue(\d+)', color_def)
        if r_match and g_match and b_match:
            r_val, g_val, b_val = int(r_match.group(1)), int(g_match.group(1)), int(b_match.group(1))
            if r_val == 0 and g_val == 128 and b_val == 0:
                green_cf = f'\\cf{idx}'
            elif r_val == 0 and g_val == 0 and b_val == 0:
                black_cf = f'\\cf{idx}'
            elif r_val == 255 and g_val == 0 and b_val == 0:
                red_cf = f'\\cf{idx}'
    return green_cf or r'\cf2', black_cf or r'\cf1', red_cf


# Cấu hình log và giữ lại 3 ngày gần nhất
LOG_FILE = "app_bien_muc_tg24h.log"

def prune_log_file(log_path, days=3):
    if not os.path.exists(log_path):
        return
    cutoff = datetime.datetime.now() - datetime.timedelta(days=days)
    try:
        valid_lines = []
        with open(log_path, "r", encoding="utf-8") as f:
            for line in f:
                # Mỗi dòng log bắt đầu bằng: YYYY-MM-DD
                match = re.match(r"^(\d{4}-\d{2}-\d{2})", line)
                if match:
                    try:
                        log_date = datetime.datetime.strptime(match.group(1), "%Y-%m-%d")
                        if log_date >= cutoff:
                            valid_lines.append(line)
                    except ValueError:
                        valid_lines.append(line)
                else:
                    if valid_lines:
                        valid_lines.append(line)
        with open(log_path, "w", encoding="utf-8") as f:
            f.writelines(valid_lines)
    except Exception as e:
        print(f"Lỗi khi dọn dẹp log cũ: {e}")

prune_log_file(LOG_FILE, 3)

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8"
)

CONFIG_FILE = "config.json"

# --- Định nghĩa Schema cho Gemini API ---
class NewsItem(typing.TypedDict):
    id: str
    ten_file: str
    thoi_luong: str

class ListParseResult(typing.TypedDict):
    ngay_phat_song: str
    danh_sach_tin: list[NewsItem]

class CrewList(typing.TypedDict):
    chiu_trach_nhiem: str
    bien_tap: str
    bien_dich: str
    hien_dan: str
    dao_dien: str
    ky_thuat: str
    trang_diem: str

class RtfParseResult(typing.TypedDict):
    tieu_de: str
    nguoi_bien_dich: str
    noi_dung: list[str]

# --- App Chính ---
class BienMucApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Tự Động Biên Mục Thế Giới 24H (TG24H)")
        self.root.geometry("800x600")
        self.root.resizable(True, True)

        self.input_dir = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.api_key = tk.StringVar()
        self.model_name = tk.StringVar(value="gemini-1.5-flash") # Mặc định
        self.fallback_model_1 = tk.StringVar(value="gemini-1.5-pro") # Mặc định
        self.fallback_model_2 = tk.StringVar(value="gemini-2.0-flash") # Mặc định

        # --- Provider 2: Mistral ---
        self.provider = tk.StringVar(value="gemini")  # "gemini", "mistral" hoặc "openai_compat"
        self.mistral_api_key = tk.StringVar()
        self.mistral_model_name = tk.StringVar(value="mistral-medium-latest")
        self.mistral_fallback_1 = tk.StringVar(value="mistral-small-latest")
        self.mistral_fallback_2 = tk.StringVar(value="mistral-small-2409")

        # --- Provider 3: OpenAI-compatible ---
        self.openai_compat_api_key = tk.StringVar()
        self.openai_compat_base_url = tk.StringVar(value="https://api.groq.com/openai/v1")
        self.openai_compat_model_name = tk.StringVar(value="openai/gpt-oss-120b")
        self.openai_compat_fallback_1 = tk.StringVar(value="llama-3.3-70b-versatile")
        self.openai_compat_fallback_2 = tk.StringVar(value="qwen/qwen3-32b")
        
        self.load_config()
        self.build_ui()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    self.api_key.set(cfg.get("api_key", ""))
                    self.model_name.set(cfg.get("model_name", "gemini-1.5-flash"))
                    self.fallback_model_1.set(cfg.get("fallback_model_1", "gemini-1.5-pro"))
                    self.fallback_model_2.set(cfg.get("fallback_model_2", "gemini-2.0-flash"))
                    self.provider.set(cfg.get("provider", "gemini"))
                    self.mistral_api_key.set(cfg.get("mistral_api_key", ""))
                    self.mistral_model_name.set(cfg.get("mistral_model_name", "mistral-medium-latest"))
                    self.mistral_fallback_1.set(cfg.get("mistral_fallback_1", "mistral-small-latest"))
                    self.mistral_fallback_2.set(cfg.get("mistral_fallback_2", "mistral-small-2409"))
                    self.openai_compat_api_key.set(cfg.get("openai_compat_api_key", ""))
                    self.openai_compat_base_url.set(cfg.get("openai_compat_base_url", "https://api.groq.com/openai/v1"))
                    self.openai_compat_model_name.set(cfg.get("openai_compat_model_name", "openai/gpt-oss-120b"))
                    self.openai_compat_fallback_1.set(cfg.get("openai_compat_fallback_1", "llama-3.3-70b-versatile"))
                    self.openai_compat_fallback_2.set(cfg.get("openai_compat_fallback_2", "qwen/qwen3-32b"))
            except Exception as e:
                logging.error(f"Lỗi đọc config: {e}")
        
        if not self.api_key.get() and "GEMINI_API_KEY" in os.environ:
            self.api_key.set(os.environ["GEMINI_API_KEY"])
        if not self.mistral_api_key.get() and "MISTRAL_API_KEY" in os.environ:
            self.mistral_api_key.set(os.environ["MISTRAL_API_KEY"])

    def save_config(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({
                    "api_key": self.api_key.get(),
                    "model_name": self.model_name.get(),
                    "fallback_model_1": self.fallback_model_1.get(),
                    "fallback_model_2": self.fallback_model_2.get(),
                    "provider": self.provider.get(),
                    "mistral_api_key": self.mistral_api_key.get(),
                    "mistral_model_name": self.mistral_model_name.get(),
                    "mistral_fallback_1": self.mistral_fallback_1.get(),
                    "mistral_fallback_2": self.mistral_fallback_2.get(),
                    "openai_compat_api_key": self.openai_compat_api_key.get(),
                    "openai_compat_base_url": self.openai_compat_base_url.get(),
                    "openai_compat_model_name": self.openai_compat_model_name.get(),
                    "openai_compat_fallback_1": self.openai_compat_fallback_1.get(),
                    "openai_compat_fallback_2": self.openai_compat_fallback_2.get()
                }, f, indent=4)
        except Exception as e:
            logging.error(f"Lỗi lưu config: {e}")

    def build_ui(self):
        # Frame trên: Chọn folder & Cấu hình mã bản tin
        frame_top = ttk.LabelFrame(self.root, text="Cấu hình Đường dẫn & Bản tin", padding=(10, 10))
        frame_top.pack(fill="x", padx=10, pady=5)

        ttk.Label(frame_top, text="Thư mục Input:").grid(row=0, column=0, sticky="w", pady=5)
        ttk.Entry(frame_top, textvariable=self.input_dir, width=60).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame_top, text="Chọn...", command=self.browse_input).grid(row=0, column=2, pady=5)

        ttk.Label(frame_top, text="Thư mục Output:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(frame_top, textvariable=self.output_dir, width=60).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(frame_top, text="Chọn...", command=self.browse_output).grid(row=1, column=2, pady=5)
        ttk.Label(frame_top, text="(Mặc định tạo folder 'output' trong input nếu để trống)", font=("Arial", 8, "italic")).grid(row=2, column=1, sticky="w")

        ttk.Label(frame_top, text="Mã bản tin ($a090):").grid(row=3, column=0, sticky="w", pady=5)
        self.a090_entry = ttk.Entry(frame_top, width=20)
        self.a090_entry.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        
        # Thiết lập chữ gợi ý mờ (Placeholder)
        def set_placeholder(entry, text):
            entry.insert(0, text)
            entry.config(foreground='grey')

        def on_entry_click(event, entry, text):
            if entry.get() == text:
                entry.delete(0, tk.END)
                entry.config(foreground='black')

        def on_focusout(event, entry, text):
            if entry.get() == '':
                set_placeholder(entry, text)

        set_placeholder(self.a090_entry, "K303419")
        self.a090_entry.bind('<FocusIn>', lambda event: on_entry_click(event, self.a090_entry, "K303419"))
        self.a090_entry.bind('<FocusOut>', lambda event: on_focusout(event, self.a090_entry, "K303419"))

        # Frame giữa: Controls
        frame_mid = ttk.Frame(self.root, padding=(10, 5))
        frame_mid.pack(fill="x", padx=10)

        ttk.Button(frame_mid, text="⚙ Cài đặt API", command=self.open_settings).pack(side="left")
        
        self.btn_start = ttk.Button(frame_mid, text="▶ BẮT ĐẦU BIÊN MỤC", command=self.start_process, style="Accent.TButton")
        self.btn_start.pack(side="right")

        # Khung lưu ý định dạng đầu vào
        frame_note = ttk.LabelFrame(self.root, text="Lưu ý định dạng đầu vào để bóc tách đúng", padding=(10, 5))
        frame_note.pack(fill="x", padx=10, pady=5)
        note_text = (
            "• File LIST (Excel): Bắt đầu bằng 'BTTG24H_' (.xlsx).\n"
            "   Cột A: Tên file ('24H-' hoặc 'GAT24H-'), Cột C: ID.\n"
            ".\n"
        )
        ttk.Label(frame_note, text=note_text, justify="left", font=("Arial", 9)).pack(anchor="w")

        # Progress
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.root, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill="x", padx=10, pady=5)

        # Log Text
        self.txt_log = scrolledtext.ScrolledText(self.root, state="disabled", wrap="word", font=("Consolas", 10))
        self.txt_log.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def browse_input(self):
        d = filedialog.askdirectory(title="Chọn thư mục Input chứa file XLSX và RTF")
        if d: self.input_dir.set(d)

    def browse_output(self):
        d = filedialog.askdirectory(title="Chọn thư mục Output")
        if d: self.output_dir.set(d)

    def log(self, message, to_gui=True):
        logging.info(message)
        if to_gui:
            self.root.after(0, self._append_log, message)

    def _append_log(self, message):
        self.txt_log.config(state="normal")
        self.txt_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state="disabled")

    def set_progress(self, val):
        self.root.after(0, self._set_progress_gui, val)

    def _set_progress_gui(self, val):
        self.progress_var.set(val)
        self.root.update_idletasks()

    def _get_provider_display_name(self, provider: str) -> str:
        if provider.lower() == "gemini":
            return "Provider 1"
        elif provider.lower() == "mistral":
            return "Provider 2"
        elif provider.lower() == "openai_compat":
            return "Provider 3"
        return provider

    def _get_model_display_name(self, m_name: str) -> str:
        m_name = m_name.strip()
        if m_name == self.model_name.get().strip() or m_name == self.mistral_model_name.get().strip() or m_name == self.openai_compat_model_name.get().strip():
            return "Primary"
        elif m_name == self.fallback_model_1.get().strip() or m_name == self.mistral_fallback_1.get().strip() or m_name == self.openai_compat_fallback_1.get().strip():
            return "Fallback 1"
        elif m_name == self.fallback_model_2.get().strip() or m_name == self.mistral_fallback_2.get().strip() or m_name == self.openai_compat_fallback_2.get().strip():
            return "Fallback 2"
        return m_name

    def open_settings(self):
        top = tk.Toplevel(self.root)
        top.title("Cài đặt API")
        top.geometry("560x460")
        top.resizable(False, False)
        top.transient(self.root)
        top.grab_set()

        # --- Chọn Provider ---
        frm_provider = ttk.LabelFrame(top, text="Nhà cung cấp dịch vụ (Provider)", padding=(10, 6))
        frm_provider.pack(fill="x", padx=10, pady=(10, 5))
        ttk.Radiobutton(frm_provider, text="Provider 1 (mặc định)",
                        variable=self.provider, value="gemini",
                        command=lambda: self._toggle_provider_frames(frm_gemini, frm_mistral, frm_openai_compat)).pack(anchor="w")
        ttk.Radiobutton(frm_provider, text="Provider 2",
                        variable=self.provider, value="mistral",
                        command=lambda: self._toggle_provider_frames(frm_gemini, frm_mistral, frm_openai_compat)).pack(anchor="w")
        ttk.Radiobutton(frm_provider, text="Provider 3 (OpenAI-compatible)",
                        variable=self.provider, value="openai_compat",
                        command=lambda: self._toggle_provider_frames(frm_gemini, frm_mistral, frm_openai_compat)).pack(anchor="w")

        # --- Frame Gemini ---
        frm_gemini = ttk.LabelFrame(top, text="Cấu hình Provider 1", padding=(10, 6))
        frm_gemini.pack(fill="x", padx=10, pady=5)

        ttk.Label(frm_gemini, text="API Key Provider 1:").pack(anchor="w")
        ent_key = ttk.Entry(frm_gemini, textvariable=self.api_key, show="*")
        ent_key.pack(fill="x", pady=(0, 5))

        # --- Frame Mistral ---
        frm_mistral = ttk.LabelFrame(top, text="Cấu hình Provider 2", padding=(10, 6))
        frm_mistral.pack(fill="x", padx=10, pady=5)

        ttk.Label(frm_mistral, text="API Key Provider 2:").pack(anchor="w")
        ttk.Entry(frm_mistral, textvariable=self.mistral_api_key, show="*").pack(fill="x", pady=(0, 5))

        # --- Frame OpenAI-compatible ---
        frm_openai_compat = ttk.LabelFrame(top, text="Cấu hình Provider 3 (OpenAI-compatible)", padding=(10, 6))
        frm_openai_compat.pack(fill="x", padx=10, pady=5)

        ttk.Label(frm_openai_compat, text="API Key Provider 3:").pack(anchor="w")
        ttk.Entry(frm_openai_compat, textvariable=self.openai_compat_api_key, show="*").pack(fill="x", pady=(0, 5))
        ttk.Label(frm_openai_compat, text="Base URL (ví dụ: https://api.groq.com/openai/v1):").pack(anchor="w")
        ttk.Entry(frm_openai_compat, textvariable=self.openai_compat_base_url).pack(fill="x", pady=(0, 5))

        # Áp dụng trạng thái ẩn/hiện ban đầu
        self._toggle_provider_frames(frm_gemini, frm_mistral, frm_openai_compat)

        def save():
            self.save_config()
            self.log(f"Đã lưu cấu hình – Provider: {self._get_provider_display_name(self.provider.get())}.")
            top.destroy()

        ttk.Button(top, text="Lưu & Đóng", command=save).pack(pady=12)

    def _toggle_provider_frames(self, frm_gemini, frm_mistral, frm_openai_compat):
        """Làm nổi bật frame của provider đang chọn, làm mờ frame còn lại."""
        current = self.provider.get()
        frames = {
            "gemini": frm_gemini,
            "mistral": frm_mistral,
            "openai_compat": frm_openai_compat,
        }
        for prov, frm in frames.items():
            state = "normal" if prov == current else "disabled"
            for w in frm.winfo_children():
                try:
                    w.configure(state=state)
                except:
                    pass

    def start_process(self):
        input_d = self.input_dir.get().strip()
        if not input_d or not os.path.isdir(input_d):
            messagebox.showerror("Lỗi", "Vui lòng chọn thư mục Input hợp lệ!")
            return

        a090_val = self.a090_entry.get().strip()
        if not a090_val or (a090_val == "K303419" and str(self.a090_entry.cget("foreground")) == "grey"):
            messagebox.showerror("Lỗi", "Vui lòng nhập mã bản tin ($a090)!")
            return

        provider = self.provider.get()
        if provider == "gemini" and not self.api_key.get().strip():
            messagebox.showerror("Lỗi", "Vui lòng nhập API Key Provider 1 trong phần Cài đặt!")
            return
        if provider == "mistral" and not self.mistral_api_key.get().strip():
            messagebox.showerror("Lỗi", "Vui lòng nhập API Key Provider 2 trong phần Cài đặt!")
            return
        if provider == "openai_compat" and not self.openai_compat_api_key.get().strip():
            messagebox.showerror("Lỗi", "Vui lòng nhập API Key Provider 3 trong phần Cài đặt!")
            return
        if provider == "openai_compat" and not self.openai_compat_base_url.get().strip():
            messagebox.showerror("Lỗi", "Vui lòng nhập Base URL Provider 3 trong phần Cài đặt!")
            return

        self.btn_start.config(state="disabled")
        self.set_progress(0)
        self.txt_log.config(state="normal")
        self.txt_log.delete(1.0, tk.END)
        self.txt_log.config(state="disabled")

        thread = threading.Thread(target=self.process_thread)
        thread.daemon = True
        thread.start()

    # ------------------------------------------------------------------
    # Helper: Bóc tách thông tin ê-kíp sản xuất bằng thuật toán nội bộ (không AI)
    # ------------------------------------------------------------------
    def _internal_extract_crew(self, ekip_text: str) -> dict:
        crew = {
            "chiu_trach_nhiem": "",
            "bien_tap": "",
            "bien_dich": "",
            "hien_dan": "",
            "dao_dien": "",
            "ky_thuat": "",
            "trang_diem": ""
        }
        patterns = {
            "chiu_trach_nhiem": [
                r"CHỊU\s+TRÁCH\s+NHIỆM\s+NỘI\s+DUNG\s*:\s*(.*)",
                r"CHIU\s+TRACH\s+NHIEM\s+NOI\s+DUNG\s*:\s*(.*)"
            ],
            "bien_tap": [
                r"BIÊN\s+TẬP\s*:\s*(.*)",
                r"BIEN\s+TAP\s*:\s*(.*)"
            ],
            "bien_dich": [
                r"BIÊN\s+DỊCH\s*:\s*(.*)",
                r"BIEN\s+DICH\s*:\s*(.*)"
            ],
            "hien_dan": [
                r"HIỆN\s+DẪN\s*:\s*(.*)",
                r"HIEN\s+DAN\s*:\s*(.*)"
            ],
            "dao_dien": [
                r"ĐẠO\s+DIỄN\s*:\s*(.*)",
                r"DAO\s+DIEN\s*:\s*(.*)"
            ],
            "ky_thuat": [
                r"KỸ\s+THUẬT\s*:\s*(.*)",
                r"KY\s+THUAT\s*:\s*(.*)"
            ],
            "trang_diem": [
                r"TRANG\s+ĐI[ỂÊêể\s\u0309]+M\s*:\s*(.*)",
                r"TRANG\s+DI[EÊeê\s]+M\s*:\s*(.*)"
            ]
        }
        for line in ekip_text.split("\n"):
            line = line.strip()
            if not line:
                continue
            for key, regexes in patterns.items():
                matched = False
                for regex in regexes:
                    match = re.match(regex, line, re.IGNORECASE)
                    if match:
                        crew[key] = match.group(1).strip()
                        matched = True
                        break
                if matched:
                    break
        return crew

    # ------------------------------------------------------------------
    # Helper: Bóc tách nội dung RTF bằng thuật toán nội bộ (không AI)
    # Trả về (title, translator, content_lines)
    # ------------------------------------------------------------------
    def _internal_extract_content(self, rtf_raw: str, is_live: bool):
        """Thuật toán nội bộ bóc tách tiêu đề + người biên dịch + nội dung từ RTF.
        
        Quy tắc:
        - Tiêu đề: dòng đầu tiên in đậm + in HOA + màu xanh lá (trừ GẠT)
        - Người biên dịch: dòng chữ không có số/kí tự đặc biệt ngay trước tiêu đề
        - Nội dung (thường): từ dưới tiêu đề đến chữ màu đen cuối cùng
        - Nội dung (LIVE): từ dưới tiêu đề đến chữ màu đỏ cuối cùng
          (bỏ qua chữ xanh lá không in đậm)
        """
        green_cf, black_cf, red_cf = get_color_tags(rtf_raw)
        
        # Lấy RTF header để wrap paragraph
        header_match = re.search(r'\\viewkind\d\\uc\d\s*', rtf_raw)
        rtf_header = rtf_raw[:header_match.end()] if header_match else r"{\rtf1\ansi "
        
        def para_to_text(para_raw):
            try:
                wrapped = rtf_header + para_raw + r"}"
                txt = rtf_to_text(wrapped).strip()
                return re.sub(r'\s+', ' ', txt)
            except:
                return ""
        
        # Split thành paragraphs
        paragraphs = re.split(r'\\par(?![a-zA-Z])', rtf_raw)
        
        parsed_paras = []
        for idx, p in enumerate(paragraphs):
            p_clean = p.strip()
            if not p_clean:
                continue
            if idx == 0:
                pard_idx = p_clean.find(r'\pard')
                if pard_idx != -1:
                    p_clean = p_clean[pard_idx:]
            
            has_green = green_cf in p_clean
            has_black = (black_cf in p_clean) if black_cf else False
            has_red = (red_cf in p_clean) if red_cf else False
            # Nếu đoạn không có tag màu nào → thừa kế màu đen từ đoạn trước
            if not has_green and not has_black and not has_red:
                has_black = True
            is_bold = r'\b' in p_clean and (r'\b0' not in p_clean or p_clean.index(r'\b') < p_clean.index(r'\b0'))
            
            txt = para_to_text(p_clean)
            if txt:
                parsed_paras.append({
                    'text': txt,
                    'has_green': has_green,
                    'has_black': has_black,
                    'has_red': has_red,
                    'is_bold': is_bold
                })
        
        # Bước 1: Tìm tiêu đề
        title = ""
        title_para_idx = -1
        for i, p in enumerate(parsed_paras):
            if p['has_green'] and p['is_bold'] and p['text'].isupper() and len(p['text']) > 3:
                txt_upper = p['text'].upper()
                if any(k in txt_upper for k in ["GẠT TG", "GAT TG", "GAT24", "GẠT24", "HEADLINES"]):
                    continue
                title = p['text']
                title_para_idx = i
                break
        
        if title_para_idx == -1:
            # Fallback: tìm dòng viết hoa đầu tiên dài > 10
            for i, p in enumerate(parsed_paras):
                txt_upper = p['text'].upper()
                if p['text'].isupper() and len(p['text']) > 10:
                    if any(k in txt_upper for k in ["GẠT TG", "GAT TG", "GAT24", "GẠT24", "HEADLINES"]):
                        continue
                    title = p['text']
                    title_para_idx = i
                    break
        
        if title_para_idx == -1:
            return title, "", []
        
        # Bước 1b: Tìm người biên dịch
        translator = ""
        plain_text = rtf_to_text(rtf_raw)
        lines = [line.strip('\r\n') for line in plain_text.split('\n')]
        line_idx = -1
        for li, line in enumerate(lines):
            if title.strip() in line:
                line_idx = li
                break
        
        if line_idx > 0:
            for offset in range(1, 8):
                if line_idx - offset >= 0:
                    candidate = lines[line_idx - offset]
                    cc = candidate.strip()
                    if cc:
                        if not re.search(r'\d', cc) and not any(k in cc.lower() for k in ["hình", "ngày", "reuters", "cnn", "ap", "afp", "ttx", "interplay", "search", "http", "www"]):
                            if 2 <= len(cc) <= 30:
                                translator = candidate
                                break
        
        # Bước 2: Tìm phạm vi nội dung
        content_paras = parsed_paras[title_para_idx + 1:]
        
        # Tìm paragraph cuối cùng của loại màu mục tiêu
        last_content_idx = -1
        for i, p in enumerate(content_paras):
            # Dừng tại dấu phân cách == hoặc ===...
            if re.search(r'={2,}', p['text']):
                break
            if is_live:
                if p['has_red'] or p['has_black']:
                    last_content_idx = i
            else:
                if p['has_black']:
                    last_content_idx = i
        
        if last_content_idx == -1:
            return title, translator, []
        
        # Trích xuất nội dung
        content_lines = []
        for p in content_paras[:last_content_idx + 1]:
            # Với LIVE: bỏ qua chữ xanh lá không in đậm, NGOẠI TRỪ dòng viết HOA (phụ đề)
            if is_live and p['has_green'] and not p['is_bold']:
                if not (p['text'].isupper() and len(p['text']) > 3):
                    continue
            # Bỏ dòng trùng tiêu đề
            if p['text'].strip().lower() == title.strip().lower():
                continue
            # Dừng tại separator
            if re.search(r'={2,}', p['text']):
                break
            if len(p['text']) > 2:
                content_lines.append(p['text'])
        
        return title, translator, content_lines

    # ------------------------------------------------------------------
    # Helper: Gọi AI theo provider đang chọn
    # Trả về dict JSON đã parse, hoặc raise Exception nếu thất bại
    # ------------------------------------------------------------------
    def _call_ai(self, prompt: str, response_schema, models_to_try: list) -> dict:
        """Gọi provider AI (Gemini, Mistral hoặc OpenAI-compatible) lần lượt qua danh sách models.
        Trả về dict đã parse JSON. Raise Exception nếu tất cả model thất bại."""
        provider = self.provider.get()
        last_exc = None

        for m_idx, m_name in enumerate(models_to_try):
            self.log(f"  Thử model: {self._get_model_display_name(m_name)} ({self._get_provider_display_name(provider)})...")
            try:
                if provider == "gemini":
                    model_obj = genai.GenerativeModel(m_name)
                    res = model_obj.generate_content(
                        prompt,
                        generation_config=genai.GenerationConfig(
                            response_mime_type="application/json",
                            response_schema=response_schema
                        ),
                        request_options={"timeout": 240}
                    )
                    result = json.loads(res.text)
                elif provider == "mistral":  # mistral
                    MISTRAL_BASE_URL = "https://api.mistral.ai/v1"
                    client = OpenAI(
                        api_key=self.mistral_api_key.get().strip(),
                        base_url=MISTRAL_BASE_URL
                    )
                    refined_prompt = prompt
                    if "json" not in prompt.lower():
                        refined_prompt += "\nTrả về kết quả dưới dạng JSON."
                    response_format = {"type": "json_object"}
                    chat_res = client.chat.completions.create(
                        model=m_name,
                        messages=[{"role": "user", "content": refined_prompt}],
                        response_format=response_format,
                        timeout=240
                    )
                    result = json.loads(chat_res.choices[0].message.content)
                else:  # openai_compat
                    base_url = self.openai_compat_base_url.get().strip()
                    client = OpenAI(
                        api_key=self.openai_compat_api_key.get().strip(),
                        base_url=base_url
                    )
                    refined_prompt = prompt
                    if "json" not in prompt.lower():
                        refined_prompt += "\nTrả về kết quả dưới dạng JSON."
                    response_format = {"type": "json_object"}
                    chat_res = client.chat.completions.create(
                        model=m_name,
                        messages=[{"role": "user", "content": refined_prompt}],
                        response_format=response_format,
                        timeout=240
                    )
                    result = json.loads(chat_res.choices[0].message.content)

                return result

            except Exception as ex:
                self.log(f"  ⚠ Lỗi model {self._get_model_display_name(m_name)}: {ex}")
                last_exc = ex
                if m_idx < len(models_to_try) - 1:
                    time.sleep(1)

        raise Exception(f"Tất cả model thất bại. Lỗi cuối: {last_exc}")

    def process_thread(self):
        try:
            self.log("=== BẮT ĐẦU TIẾN TRÌNH ===")
            provider = self.provider.get()
            self.log(f"Provider đang dùng: {self._get_provider_display_name(provider)}")

            if provider == "gemini":
                genai.configure(api_key=self.api_key.get().strip())
            # Mistral và OpenAI-compatible dùng OpenAI client, không cần configure global

            if provider == "gemini":
                model_name = self.model_name.get().strip()
            elif provider == "mistral":
                model_name = self.mistral_model_name.get().strip()
            else:  # openai_compat
                model_name = self.openai_compat_model_name.get().strip()
            
            # Lấy mã bản tin từ giao diện
            a090_val = self.a090_entry.get().strip()
            if not a090_val:
                a090_val = "K303419"
            
            self.log(f"Đang sử dụng model: {self._get_model_display_name(model_name)}")
            
            input_d = self.input_dir.get().strip()
            output_d = self.output_dir.get().strip()
            if not output_d:
                output_d = os.path.join(input_d, "output")
            os.makedirs(output_d, exist_ok=True)
            self.log(f"Thư mục Output: {output_d}")

            # Khởi tạo và dọn dẹp thư mục tempbienmuc cùng cấp với input
            temp_d = os.path.join(os.path.dirname(input_d), "tempbienmuc")
            os.makedirs(temp_d, exist_ok=True)
            self.log(f"Đang dọn dẹp thư mục tạm: {temp_d}...")
            for f_name in os.listdir(temp_d):
                f_path = os.path.join(temp_d, f_name)
                if os.path.isfile(f_path):
                    try:
                        os.remove(f_path)
                    except Exception as ex:
                        self.log(f"  ⚠ Không thể xóa file tạm {f_name}: {ex}")

            # 1. Tìm file LIST
            list_files = [f for f in os.listdir(input_d) if f.startswith("BTTG24H_") and f.endswith(".xlsx")]
            if not list_files:
                raise Exception("Không tìm thấy file danh sách bắt đầu bằng 'BTTG24H_' và đuôi '.xlsx' trong thư mục Input.")
            list_file_path = os.path.join(input_d, list_files[0])
            self.log(f"Đã tìm thấy file danh sách: {list_files[0]}")

            # Đọc file LIST ra dạng text thô để gửi cho Gemini
            self.set_progress(10)
            self.log("Đọc dữ liệu file LIST...")
            wb_list = openpyxl.load_workbook(list_file_path, data_only=True)
            ws_list = wb_list.active
            
            list_text_data = []
            id_to_time = {}
            for row in ws_list.iter_rows(values_only=False):
                # Format time correctly from column F
                val_c = str(row[2].value).strip() if row[2].value else ""
                val_f = row[5].value
                if val_c and len(val_c) == 9:
                    if isinstance(val_f, datetime.time):
                        # Excel lưu thời lượng mà thực chất là mm:ss dưới dạng HH:MM
                        # (ví dụ 01:08 = 1 phút 8 giây → datetime.time(1, 8, 0))
                        # Nên: phút thực = .hour, giây thực = .minute
                        total_secs = val_f.hour * 60 + val_f.minute
                        mins = total_secs // 60
                        secs = total_secs % 60
                        id_to_time[val_c] = f"00:{mins:02d}:{secs:02d}"
                    elif isinstance(val_f, str):
                        tl_str = val_f.replace(" AM", "").replace(" PM", "").strip()
                        parts = tl_str.split(":")
                        if len(parts) >= 2:
                            try:
                                # Trường hợp string cũng có thể là "H:MM" = phút:giây
                                total_secs = int(parts[0]) * 60 + int(parts[1])
                                id_to_time[val_c] = f"00:{(total_secs // 60):02d}:{(total_secs % 60):02d}"
                            except:
                                pass

            # 2. Xử lý LIST bằng thuật toán nội bộ thay vì AI
            self.log("Bóc tách danh sách tin chính từ file LIST (sử dụng thuật toán nội bộ thay vì AI để tránh lỗi 504 Timeout)...", to_gui=False)
            danh_sach_tin = []
            ngay_phat = ""
            
            for row_idx, row in enumerate(ws_list.iter_rows(values_only=False), start=1):
                val_a = str(row[0].value).strip() if row[0].value else ""
                val_c = str(row[2].value).strip() if row[2].value else ""
                val_d = str(row[3].value).strip() if row[3].value else ""
                
                if row_idx <= 5 and not ngay_phat:
                    for cell in row:
                        c_val = str(cell.value) if cell.value else ""
                        match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', c_val)
                        if match:
                            d, m, y = match.groups()
                            ngay_phat = f"{y}{m.zfill(2)}{d.zfill(2)}"
                            break
                            
                if (val_a.startswith("24H-") or val_a.startswith("GAT24H-")) and val_d == "ONLINE" and len(val_c) == 9 and val_c.isdigit():
                    danh_sach_tin.append({
                        "id": val_c,
                        "ten_file": val_a,
                        "thoi_luong": id_to_time.get(val_c, "00:00:00")
                    })
                    
            if not ngay_phat:
                match_fn = re.search(r'\d{8}', list_files[0])
                if match_fn:
                    ngay_phat = match_fn.group(0)
                else:
                    ngay_phat = datetime.datetime.now().strftime("%Y%m%d")

            self.log(f"Đã tìm thấy {len(danh_sach_tin)} bản tin chính. Ngày phát sóng: {ngay_phat}")

            # 3. Parse NHUNG NGUOI THUC HIEN.rtf
            self.set_progress(20)
            ekip_file = os.path.join(input_d, "NHUNG NGUOI THUC HIEN.rtf")
            crew_data = {}
            if os.path.exists(ekip_file):
                self.log("Bóc tách thông tin ê-kíp sản xuất từ NHUNG NGUOI THUC HIEN.rtf...", to_gui=False)
                with open(ekip_file, 'rb') as f:
                    rtf_raw = f.read().decode('utf-8', errors='replace')
                ekip_text = rtf_to_text(rtf_raw)
                
                prompt_crew = f"""
Trích xuất thông tin người đảm nhận các chức danh từ văn bản ê-kíp chương trình. Nếu chức danh không có tên người, trả về chuỗi rỗng "". Tên người VIẾT HOA.
Văn bản:
{ekip_text}
"""
                provider = self.provider.get()
                if provider == "mistral":
                    models_to_try = [
                        self.mistral_model_name.get().strip(),
                        self.mistral_fallback_1.get().strip(),
                        self.mistral_fallback_2.get().strip()
                    ]
                elif provider == "openai_compat":
                    models_to_try = [
                        self.openai_compat_model_name.get().strip(),
                        self.openai_compat_fallback_1.get().strip(),
                        self.openai_compat_fallback_2.get().strip()
                    ]
                else:
                    models_to_try = [
                        self.model_name.get().strip(),
                        self.fallback_model_1.get().strip(),
                        self.fallback_model_2.get().strip()
                    ]
                models_to_try = [m for m in models_to_try if m]

                try:
                    crew_data = self._call_ai(prompt_crew, CrewList, models_to_try)
                    self.log("  ✓ Bóc tách ê-kíp thành công.", to_gui=False)
                except Exception as e:
                    self.log(f"  ⚠ CẢNH BÁO: Tất cả các model đều thất bại khi bóc tách ê-kíp: {e}")
                    crew_data = {k: "" for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat", "trang_diem"]}
            else:
                self.log("Không tìm thấy file NHUNG NGUOI THUC HIEN.rtf, bỏ qua ê-kíp.")
                crew_data = {k: "" for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat", "trang_diem"]}

            # Fallback: Bổ sung từ các file RTF tiền tố nếu crew_data còn thiếu
            PREFIX_MAP = {
                "BGĐ ": "chiu_trach_nhiem",
                "BT ":  "bien_tap",
                "BD ":  "bien_dich",
                "MC ":  "hien_dan",
                "ĐD ":  "dao_dien",
                "KT ":  "ky_thuat",
            }
            # Kiểm tra xem có trường nào bị rỗng không
            missing_keys = [k for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat"] if not crew_data.get(k, "").strip()]
            if missing_keys:
                self.log("Một số chức danh còn thiếu tên — đang tìm file RTF tiền tố bổ sung...", to_gui=False)
                for fname in os.listdir(input_d):
                    if not fname.endswith(".rtf"):
                        continue
                    for prefix, field_key in PREFIX_MAP.items():
                        if fname.startswith(prefix) and field_key in missing_keys:
                            name_part = fname[len(prefix):][:-4].strip()
                            name_part = re.sub(r"\s*,\s*", " - ", name_part)
                            crew_data[field_key] = name_part.upper()
                            missing_keys.remove(field_key)
                            self.log(f"  Bổ sung từ '{fname}': {field_key} = {crew_data[field_key]}", to_gui=False)
                            break

            # Bóc tách ê-kíp sản xuất bằng thuật toán nội bộ (không AI)
            crew_data_internal = {k: "" for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat", "trang_diem"]}
            if os.path.exists(ekip_file):
                self.log("Bóc tách thông tin ê-kíp sản xuất bằng thuật toán nội bộ...", to_gui=False)
                try:
                    crew_data_internal = self._internal_extract_crew(ekip_text)
                    self.log("  ✓ Bóc tách ê-kíp bằng thuật toán nội bộ thành công.", to_gui=False)
                except Exception as e:
                    self.log(f"  ⚠ Lỗi khi bóc tách ê-kíp bằng thuật toán nội bộ: {e}")

            # Fallback cho ê-kíp nội bộ
            missing_keys_internal = [k for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat"] if not crew_data_internal.get(k, "").strip()]
            if missing_keys_internal:
                self.log("Một số chức danh nội bộ còn thiếu tên — đang tìm file RTF tiền tố bổ sung...", to_gui=False)
                for fname in os.listdir(input_d):
                    if not fname.endswith(".rtf"):
                        continue
                    for prefix, field_key in PREFIX_MAP.items():
                        if fname.startswith(prefix) and field_key in missing_keys_internal:
                            name_part = fname[len(prefix):][:-4].strip()
                            name_part = re.sub(r"\s*,\s*", " - ", name_part)
                            crew_data_internal[field_key] = name_part.upper()
                            missing_keys_internal.remove(field_key)
                            self.log(f"  Bổ sung (nội bộ) từ '{fname}': {field_key} = {crew_data_internal[field_key]}", to_gui=False)
                            break

            self.log("Đã bóc tách xong ê-kíp sản xuất")

            # 4. Parse từng file RTF tin chính
            self.log("Bắt đầu bóc tách nội dung chi tiết từng file kịch bản (RTF)...", to_gui=False)
            chi_tiet_tin = [] # Chứa dict: id, a245, a500, a520_list

            total_tin = len(danh_sach_tin)
            for tin_idx, tin in enumerate(danh_sach_tin):
                id_tin = tin["id"]
                ten_tin = tin["ten_file"]
                
                # Tìm file rtf tương ứng (xử lý ký tự / bị thay đổi trong tên file)
                safe_name = ten_tin.replace("/", "")
                rtf_path = None
                for fname in os.listdir(input_d):
                    if fname.endswith(".rtf") and safe_name.lower() in fname.replace("/", "").lower():
                        rtf_path = os.path.join(input_d, fname)
                        break
                
                if not rtf_path:
                    # Fallback tìm kiếm gần đúng bằng prefix
                    for fname in os.listdir(input_d):
                        if fname.endswith(".rtf") and fname.startswith(ten_tin[:15]):
                            rtf_path = os.path.join(input_d, fname)
                            break

                if not rtf_path:
                    self.log(f"CẢNH BÁO: Không tìm thấy file RTF cho tin '{ten_tin}' (ID: {id_tin})")
                    chi_tiet_tin.append({"id": id_tin, "tieu_de": ten_tin.upper(), "nguoi_bien_dich": "", "noi_dung": []})
                    continue

                self.log(f"Đang bóc tách: {os.path.basename(rtf_path)}...", to_gui=False)
                with open(rtf_path, 'rb') as f:
                    rtf_raw = f.read().decode('utf-8', errors='replace')

                # Định nghĩa header và tìm mã màu xanh lá cây
                rtf_header = r"{\rtf1\ansi "
                header_match = re.search(r'\\viewkind\d\\uc\d\s*', rtf_raw)
                if header_match:
                    rtf_header = rtf_raw[:header_match.end()]

                green_tag = get_green_cf_tag(rtf_raw)
                is_live_feed = "live" in ten_tin.lower()

                # Bóc tách tiêu đề theo định dạng: dòng đầu tiên in đậm (\b), in HOA, màu xanh lá cây
                tieu_de_tu_dinh_dang = ""
                try:
                    paragraphs = re.split(r'\\par(?![a-zA-Z])', rtf_raw)
                    for p_idx, p in enumerate(paragraphs):
                        p_clean = p.strip()
                        if not p_clean:
                            continue
                        if p_idx == 0:
                            # Tránh duplicate header làm hỏng parsing dòng đầu tiên
                            pard_idx = p_clean.find(r'\pard')
                            if pard_idx != -1:
                                p_clean = p_clean[pard_idx:]

                        is_green = green_tag in p_clean
                        is_bold = r'\b' in p_clean and (r'\b0' not in p_clean or p_clean.index(r'\b') < p_clean.index(r'\b0'))
                        if is_green and is_bold:
                            rtf_wrapped = rtf_header + p_clean + r"}"
                            txt = rtf_to_text(rtf_wrapped).strip()
                            txt = re.sub(r'\s+', ' ', txt)
                            if txt and txt.isupper() and len(txt) > 3:
                                # Bỏ qua dòng phân cách phân loại như "GẠT TG24H", "HEADLINES"...
                                txt_upper = txt.upper()
                                if any(k in txt_upper for k in ["GẠT TG", "GAT TG", "GAT24", "GẠT24", "HEADLINES"]):
                                    continue
                                tieu_de_tu_dinh_dang = txt
                                break
                    if tieu_de_tu_dinh_dang:
                        self.log(f"  ✓ Tìm thấy tiêu đề định dạng: '{tieu_de_tu_dinh_dang}'", to_gui=False)
                except Exception as e:
                    self.log(f"  ⚠ Lỗi khi trích xuất tiêu đề theo định dạng: {e}")

                # Bóc tách văn bản thô (có lọc tin LIVE nếu cần)
                if is_live_feed:
                    try:
                        paragraphs = re.split(r'\\par(?![a-zA-Z])', rtf_raw)
                        filtered_paragraphs = []
                        found_title = False
                        seen_content_after_title = False
                        for p_idx, p in enumerate(paragraphs):
                            p_clean = p.strip()
                            if not p_clean:
                                continue
                            if p_idx == 0:
                                pard_idx = p_clean.find(r'\pard')
                                if pard_idx != -1:
                                    p_clean = p_clean[pard_idx:]

                            has_green = green_tag in p_clean
                            is_bold = r'\b' in p_clean and (r'\b0' not in p_clean or p_clean.index(r'\b') < p_clean.index(r'\b0'))

                            rtf_wrapped = rtf_header + p_clean + r"}"
                            txt = rtf_to_text(rtf_wrapped).strip()
                            txt = re.sub(r'\s+', ' ', txt)

                            if not txt:
                                continue

                            # Phát hiện tiêu đề: green + bold + HOA
                            if not found_title and has_green and is_bold and txt.isupper() and len(txt) > 3:
                                txt_upper = txt.upper()
                                if not any(k in txt_upper for k in ["GẠT TG", "GAT TG", "GAT24", "GẠT24", "HEADLINES"]):
                                    found_title = True

                            # Phát hiện nội dung đỏ/đen sau tiêu đề
                            if found_title and not has_green and len(txt) > 2:
                                seen_content_after_title = True

                            # Lọc green non-bold theo vị trí:
                            if has_green and not is_bold:
                                # Giữ phụ đề HOA nằm giữa tiêu đề và nội dung đỏ/đen
                                if found_title and not seen_content_after_title and txt.isupper() and len(txt) > 3:
                                    pass  # Giữ lại (phụ đề)
                                else:
                                    continue  # Bỏ (metadata trước tiêu đề / CG text sau nội dung)

                            filtered_paragraphs.append(txt)
                        news_text = '\n'.join(filtered_paragraphs)
                    except Exception as e:
                        self.log(f"  ⚠ Lỗi khi lọc tin LIVE: {e}. Fallback sang text thô.")
                        news_text = rtf_to_text(rtf_raw)
                else:
                    news_text = rtf_to_text(rtf_raw)

                # Cắt bỏ nội dung từ dòng có 3 ký tự '=' liên tục trở lên đến cuối
                filtered_lines = []
                for line in news_text.split('\n'):
                    if re.search(r'={3,}', line):
                        break
                    filtered_lines.append(line)
                news_text = '\n'.join(filtered_lines)

                prompt_news = f"""
Trích xuất thông tin từ kịch bản bản tin sau.
1. tieu_de: Tiêu đề bản tin, thường được viết HOA toàn bộ (ví dụ: THIỆT HẠI DO ĐỘNG ĐẤT Ở PHILIPPINES TIẾP TỤC TĂNG). Nó thường nằm ở dòng thứ 1, 3, hoặc 5 của văn bản. Bỏ qua dòng chữ 'GẠT TG24H' và các tiêu đề tiếng Anh. Bắt buộc phải trích xuất được tiêu đề.
2. nguoi_bien_dich: Tên người biên dịch bản tin (tên người Việt), thường nằm trước tiêu đề chính. Nếu không có hoặc không rõ thì để chuỗi rỗng "", TUYỆT ĐỐI KHÔNG tự bịa ra tên.
3. noi_dung: Danh sách các đoạn văn bản cấu thành nội dung tin. Bao gồm các dòng phụ đề viết HOA và các đoạn lời đọc. Bỏ qua các dòng mã hình/video, bỏ qua tên tiếng Anh, bỏ qua ngày tháng. Giữ nguyên format viết hoa của phụ đề. Mỗi đoạn/câu là một phần tử trong mảng.

Văn bản:
{news_text}
"""
                provider = self.provider.get()
                if provider == "mistral":
                    models_to_try = [
                        self.mistral_model_name.get().strip(),
                        self.mistral_fallback_1.get().strip(),
                        self.mistral_fallback_2.get().strip()
                    ]
                elif provider == "openai_compat":
                    models_to_try = [
                        self.openai_compat_model_name.get().strip(),
                        self.openai_compat_fallback_1.get().strip(),
                        self.openai_compat_fallback_2.get().strip()
                    ]
                else:
                    models_to_try = [
                        self.model_name.get().strip(),
                        self.fallback_model_1.get().strip(),
                        self.fallback_model_2.get().strip()
                    ]
                models_to_try = [m for m in models_to_try if m]

                news_parsed = {}
                success_news = False
                # Thử gọi AI – nếu thành công nhưng noi_dung rỗng thì cũng xem là thất bại
                try:
                    news_parsed = self._call_ai(prompt_news, RtfParseResult, models_to_try)
                    if news_parsed.get("noi_dung") and len(news_parsed["noi_dung"]) > 0:
                        success_news = True
                        self.log(f"  ✓ Bóc tách thành công.", to_gui=False)
                        self.log(f"Bóc tách thành công {os.path.basename(rtf_path)}")
                    else:
                        self.log(f"  ⚠ AI trả về noi_dung rỗng cho {os.path.basename(rtf_path)}.", to_gui=False)
                except Exception as ex:
                    self.log(f"  ⚠ Tất cả model thất bại cho {os.path.basename(rtf_path)}: {ex}", to_gui=False)
                
                tieu_de_ai = tieu_de_tu_dinh_dang if tieu_de_tu_dinh_dang else news_parsed.get("tieu_de", "").strip()
                noi_dung_parsed = news_parsed.get("noi_dung", [])

                if not tieu_de_ai:
                    # Fallback: tìm dòng viết hoa đầu tiên dài hơn 10 ký tự, không chứa GẠT TG24H
                    for line in news_text.split('\n'):
                        line = line.strip()
                        if line.isupper() and len(line) > 10 and "GẠT TG24H" not in line and "HEADLINES" not in line:
                            tieu_de_ai = line
                            break

                # Fallback nội dung: nếu sau 3 lần AI vẫn trả noi_dung rỗng,
                # tự trích xuất toàn bộ văn bản phía dưới tiêu đề
                if not noi_dung_parsed:
                    self.log(f"  ⚠ CẢNH BÁO: AI không trả về nội dung cho '{os.path.basename(rtf_path)}' sau 3 lần thử.")
                    self.log(f"  → Chuyển sang trích xuất nội dung bằng thuật toán nội bộ (không dùng AI)...", to_gui=False)
                    all_lines = [l.strip() for l in news_text.split('\n')]
                    # Tìm vị trí tiêu đề trong văn bản
                    title_idx = -1
                    if tieu_de_ai:
                        for li, line in enumerate(all_lines):
                            if tieu_de_ai in line:
                                title_idx = li
                                break
                    if title_idx == -1:
                        # Nếu không tìm được tiêu đề, tìm dòng viết hoa đầu tiên dài > 10
                        for li, line in enumerate(all_lines):
                            if line.isupper() and len(line) > 10 and "GẠT TG24H" not in line and "HEADLINES" not in line:
                                title_idx = li
                                break
                    # Lấy tất cả đoạn phía dưới tiêu đề
                    if title_idx >= 0:
                        content_lines = all_lines[title_idx + 1:]
                    else:
                        content_lines = all_lines  # Không tìm được tiêu đề → lấy toàn bộ
                    # Lọc bỏ dòng trống và dòng quá ngắn (<=2 ký tự)
                    noi_dung_parsed = [l for l in content_lines if len(l) > 2]
                    self.log(f"  ✓ Đã trích xuất được {len(noi_dung_parsed)} đoạn nội dung bằng thuật toán nội bộ.", to_gui=False)
                    self.log(f"Bóc tách bằng thuật toán nội bộ {os.path.basename(rtf_path)}")

                # --- Bóc tách nội bộ (không AI) để so sánh cho Map_ChiTiet ---
                internal_title, internal_translator, internal_content = self._internal_extract_content(rtf_raw, is_live_feed)
                
                chi_tiet_tin.append({
                    "id": id_tin,
                    "tieu_de": tieu_de_ai,
                    "nguoi_bien_dich": news_parsed.get("nguoi_bien_dich", ""),
                    "noi_dung": noi_dung_parsed,
                    "internal_title": internal_title,
                    "internal_translator": internal_translator,
                    "internal_content": internal_content,
                    "is_live": is_live_feed
                })

                self.set_progress(20 + (tin_idx + 1) / total_tin * 50) # Cập nhật progress 20 -> 70%

            # 5. Sinh Output 1: Import_SoLuoc
            self.log("Đang tạo file Output 1: Import_SoLuoc...", to_gui=False)
            wb1 = openpyxl.Workbook()
            ws1 = wb1.active
            ws1.title = "Sheet1"
            headers1 = ["STT", "$a090", "$a245", "$n245", "$p245", "$b245", "$a246", "$a260", "$b260", "$c260", "$a300", "$c300", "$a306", "$a490", "$a500", "$t773", "$r773", "$r773", "$a911"]
            ws1.append(headers1)
            
            nam = ngay_phat[:4] if len(ngay_phat) >= 4 else str(datetime.datetime.now().year)
            thang = ngay_phat[4:6] if len(ngay_phat) >= 6 else str(datetime.datetime.now().month).zfill(2)
            ngay = ngay_phat[6:8] if len(ngay_phat) >= 8 else str(datetime.datetime.now().day).zfill(2)

            for idx, tin in enumerate(danh_sach_tin):
                # Match title
                tieu_de = ""
                for ct in chi_tiet_tin:
                    if ct["id"] == tin["id"]:
                        tieu_de = ct["tieu_de"]
                        break
                
                row_data = [
                    f"{(idx+1):02d}", # STT
                    tin["id"],        # $a090
                    tieu_de,          # $a245
                    "", "",           # $n245, $p245
                    f"Tin thế giới - bản tin 24g ngày {ngay}/{thang}/{nam}", # $b245
                    "",               # $a246
                    "Tp.HCM",         # $a260
                    "Trung tâm tin tức HTV", # $b260
                    nam,              # $c260
                    "File MXF",       # $a300
                    "",               # $c300
                    tin["thoi_luong"],# $a306
                    "",               # $a490
                    f"Tên file: {tin['id']}  {tin['ten_file']}", # $a500
                    "", "", "",       # $t773, $r773 x2
                    "Trung tâm Phát hình - Tư liệu HTV" # $a911
                ]
                ws1.append(row_data)
            
            apply_tnr_font(ws1)
            fn1 = os.path.join(output_d, f"Import_SoLuoc_TG24H_{ngay_phat}.xlsx")
            wb1.save(fn1)

            # 5b. Sinh Output 1 Thuật Toán: Import_SoLuoc_TG24H_thuattoan (lưu vào tempbienmuc)
            self.log("Đang tạo file Import_SoLuoc_TG24H_thuattoan...", to_gui=False)
            wb1_tt = openpyxl.Workbook()
            ws1_tt = wb1_tt.active
            ws1_tt.title = "Sheet1"
            ws1_tt.append(headers1)
            
            for idx, tin in enumerate(danh_sach_tin):
                # Match internal title
                tieu_de_tt = ""
                for ct in chi_tiet_tin:
                    if ct["id"] == tin["id"]:
                        tieu_de_tt = ct.get("internal_title", "")
                        break
                
                row_data_tt = [
                    f"{(idx+1):02d}", # STT
                    tin["id"],        # $a090
                    tieu_de_tt,       # $a245
                    "", "",           # $n245, $p245
                    f"Tin thế giới - bản tin 24g ngày {ngay}/{thang}/{nam}", # $b245
                    "",               # $a246
                    "Tp.HCM",         # $a260
                    "Trung tâm tin tức HTV", # $b260
                    nam,              # $c260
                    "File MXF",       # $a300
                    "",               # $c300
                    tin["thoi_luong"],# $a306
                    "",               # $a490
                    f"Tên file: {tin['id']}  {tin['ten_file']}", # $a500
                    "", "", "",       # $t773, $r773 x2
                    "Trung tâm Phát hình - Tư liệu HTV" # $a911
                ]
                ws1_tt.append(row_data_tt)
            
            apply_tnr_font(ws1_tt)
            fn1_tt = os.path.join(temp_d, f"Import_SoLuoc_TG24H_thuattoan_{ngay_phat}.xlsx")
            wb1_tt.save(fn1_tt)
            self.log(f"  ✓ Đã tạo file: {os.path.basename(fn1_tt)}", to_gui=False)

            # 6. Sinh Output 2: Map_BanTinTG
            self.set_progress(75)
            self.log("Đang tạo file Output 2: Map_BanTinTG...", to_gui=False)
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = "Sheet1"
            headers2 = ["$a090", "$a500", "$a505", "$a911"]
            ws2.append(headers2)

            # Build a500 column array
            a500_crew = [
                f"CHỊU TRÁCH NHIỆM NỘI DUNG: {crew_data.get('chiu_trach_nhiem', '')}",
                f"BIÊN TẬP: {crew_data.get('bien_tap', '')}",
                f"BIÊN DỊCH: {crew_data.get('bien_dich', '')}",
                f"HIỆN DẪN: {crew_data.get('hien_dan', '')}",
                f"ĐẠO DIỄN: {crew_data.get('dao_dien', '')}",
                f"KỸ THUẬT: {crew_data.get('ky_thuat', '')}",
                f"TRANG ĐIỂM: {crew_data.get('trang_diem', '')}",
                "Website: www.htv.com.vn/tin-tuc",
                "Fanpage: www.fb.com/htvtintuc",
                "Kênh Youtube: www.youtube.com/c/htvtintuc"
            ]

            # Lặp theo danh sách tin để đủ số dòng
            max_rows = max(len(a500_crew), len(danh_sach_tin))
            
            for i in range(max_rows):
                val_a500 = a500_crew[i] if i < len(a500_crew) else ""
                
                val_a505 = ""
                if i < len(danh_sach_tin):
                    tin = danh_sach_tin[i]
                    tieu_de = ""
                    for ct in chi_tiet_tin:
                        if ct["id"] == tin["id"]:
                            tieu_de = ct["tieu_de"]
                            break
                    val_a505 = f"{(i+1):02d} - {tieu_de}. Thời lượng: {tin['thoi_luong']}. ID: {tin['id']}"
                
                val_a911 = "Phạm Thị Đông" if i == 0 else ""
                
                ws2.append([a090_val, val_a500, val_a505, val_a911])
            
            apply_tnr_font(ws2)
            fn2 = os.path.join(output_d, f"Map_BanTinTG_24G_{ngay_phat}.xlsx")
            wb2.save(fn2)

            # 6b. Sinh Output 2 Thuật Toán: Map_BanTinTG_24G_thuattoan (lưu vào tempbienmuc)
            self.log("Đang tạo file Map_BanTinTG_24G_thuattoan...", to_gui=False)
            wb2_tt = openpyxl.Workbook()
            ws2_tt = wb2_tt.active
            ws2_tt.title = "Sheet1"
            ws2_tt.append(headers2)

            # Build a500 column array using crew_data_internal
            a500_crew_tt = [
                f"CHỊU TRÁCH NHIỆM NỘI DUNG: {crew_data_internal.get('chiu_trach_nhiem', '')}",
                f"BIÊN TẬP: {crew_data_internal.get('bien_tap', '')}",
                f"BIÊN DỊCH: {crew_data_internal.get('bien_dich', '')}",
                f"HIỆN DẪN: {crew_data_internal.get('hien_dan', '')}",
                f"ĐẠO DIỄN: {crew_data_internal.get('dao_dien', '')}",
                f"KỸ THUẬT: {crew_data_internal.get('ky_thuat', '')}",
                f"TRANG ĐIỂM: {crew_data_internal.get('trang_diem', '')}",
                "Website: www.htv.com.vn/tin-tuc",
                "Fanpage: www.fb.com/htvtintuc",
                "Kênh Youtube: www.youtube.com/c/htvtintuc"
            ]

            max_rows_tt = max(len(a500_crew_tt), len(danh_sach_tin))
            
            for i in range(max_rows_tt):
                val_a500_tt = a500_crew_tt[i] if i < len(a500_crew_tt) else ""
                
                val_a505_tt = ""
                if i < len(danh_sach_tin):
                    tin = danh_sach_tin[i]
                    tieu_de_tt = ""
                    for ct in chi_tiet_tin:
                        if ct["id"] == tin["id"]:
                            tieu_de_tt = ct.get("internal_title", "")
                            break
                    val_a505_tt = f"{(i+1):02d} - {tieu_de_tt}. Thời lượng: {tin['thoi_luong']}. ID: {tin['id']}"
                
                val_a911_tt = "Phạm Thị Đông" if i == 0 else ""
                
                ws2_tt.append([a090_val, val_a500_tt, val_a505_tt, val_a911_tt])
            
            apply_tnr_font(ws2_tt)
            fn2_tt = os.path.join(temp_d, f"Map_BanTinTG_24G_thuattoan_{ngay_phat}.xlsx")
            wb2_tt.save(fn2_tt)
            self.log(f"  ✓ Đã tạo file: {os.path.basename(fn2_tt)}", to_gui=False)

            # 7. Sinh Output 3: Map_ChiTiet (AI)
            self.set_progress(80)
            self.log("Đang tạo file Output 3: Map_ChiTiet (AI)...", to_gui=False)
            wb3 = openpyxl.Workbook()
            ws3 = wb3.active
            ws3.title = "24G"
            ws3.append(["$a090", "$a500", "$a520"])

            # Đếm số dòng mỗi tin trong file AI
            ai_line_counts = {}

            for idx, ct in enumerate(chi_tiet_tin):
                id_tin = ct["id"]
                is_live = ct.get("is_live", False)
                if is_live:
                    nguoi_bd = ct.get("internal_translator", "")
                    nd_list = ct.get("internal_content", [])
                    clean_title = ct.get("internal_title", "").strip().lower() if ct.get("internal_title") else ""
                else:
                    nguoi_bd = ct["nguoi_bien_dich"]
                    nd_list = ct["noi_dung"]
                    clean_title = ct["tieu_de"].strip().lower() if ct.get("tieu_de") else ""
                
                val_a500_first = f"Biên dịch: {nguoi_bd}" if nguoi_bd else ""
                
                # Loại bỏ dòng trùng với tiêu đề chính (tieu_de)
                filtered_nd = []
                for line in nd_list:
                    if clean_title and line.strip().lower() == clean_title:
                        continue
                    filtered_nd.append(line)

                if not filtered_nd:
                    ws3.append([id_tin, val_a500_first, ""])
                    ai_line_counts[id_tin] = 1
                else:
                    for j, line in enumerate(filtered_nd):
                        a500 = val_a500_first if j == 0 else ""
                        ws3.append([id_tin, a500, line])
                    ai_line_counts[id_tin] = len(filtered_nd)

            apply_tnr_font(ws3)
            fn3 = os.path.join(output_d, f"Map_ChiTiet_24G_{ngay_phat}.xlsx")
            wb3.save(fn3)

            # 8. Sinh Output 4: Map_ChiTiet_ThuatToan (hoàn toàn từ thuật toán nội bộ)
            self.set_progress(90)
            self.log("Đang tạo file Output 4: Map_ChiTiet_ThuatToan (thuật toán nội bộ)...", to_gui=False)
            wb4 = openpyxl.Workbook()
            ws4 = wb4.active
            ws4.title = "24G"
            ws4.append(["$a090", "$a500", "$a520"])

            # Đếm số dòng mỗi tin trong file thuật toán
            tt_line_counts = {}

            for idx, ct in enumerate(chi_tiet_tin):
                id_tin = ct["id"]
                nguoi_bd = ct.get("internal_translator", "")
                int_title = ct.get("internal_title", "")
                int_content = ct.get("internal_content", [])
                
                val_a500_first = f"Biên dịch: {nguoi_bd}" if nguoi_bd else ""
                
                # Loại bỏ dòng trùng với tiêu đề nội bộ
                filtered_int = []
                clean_int_title = int_title.strip().lower() if int_title else ""
                for line in int_content:
                    if clean_int_title and line.strip().lower() == clean_int_title:
                        continue
                    filtered_int.append(line)

                if not filtered_int:
                    ws4.append([id_tin, val_a500_first, ""])
                    tt_line_counts[id_tin] = 1
                else:
                    for j, line in enumerate(filtered_int):
                        a500 = val_a500_first if j == 0 else ""
                        ws4.append([id_tin, a500, line])
                    tt_line_counts[id_tin] = len(filtered_int)

            apply_tnr_font(ws4)
            fn4 = os.path.join(temp_d, f"Map_ChiTiet_ThuatToan_{ngay_phat}.xlsx")
            wb4.save(fn4)
            self.log(f"  ✓ Đã tạo file: {os.path.basename(fn4)}", to_gui=False)
            self.log("Đã hoàn thành tạo file")

            # --- So sánh SỐ DÒNG mỗi tin giữa Map_ChiTiet (AI) và Map_ChiTiet_ThuatToan ---
            discrepancies = []
            for ct in chi_tiet_tin:
                id_tin = ct["id"]
                ai_count = ai_line_counts.get(id_tin, 0)
                tt_count = tt_line_counts.get(id_tin, 0)
                if ai_count != tt_count:
                    discrepancies.append({"id": id_tin, "ai": ai_count, "tt": tt_count})
            
            if discrepancies:
                self.log("")
                self.log("Phát hiện số dòng không trùng khớp giữa AI và ThuatToan:")
                for d in discrepancies:
                    msg = f"Tin {d['id']}: AI={d['ai']} dòng, ThuậtToán={d['tt']} dòng"
                    self.log(f"  ▸ {msg}")
                self.log("Đề nghị kiểm tay những tin trên.")
                self.log("")
            else:
                self.log("✓ Số dòng mỗi tin trong Map_ChiTiet (AI) và Map_ChiTiet_ThuatToan trùng khớp.")

            self.set_progress(100)
            self.log("=== HOÀN TẤT BIÊN MỤC ===")
            messagebox.showinfo("Thành công", f"Đã sinh các file output thành công tại:\n{output_d}")

        except Exception as e:
            err_msg = traceback.format_exc()
            self.log(f"LỖI NGHIÊM TRỌNG: {e}\n{err_msg}")
            messagebox.showerror("Lỗi", f"Đã xảy ra lỗi, vui lòng xem log tiến trình trên màn hình!\nChi tiết: {e}")
        finally:
            self.btn_start.config(state="normal")

if __name__ == "__main__":
    root = tk.Tk()
    app = BienMucApp(root)
    root.mainloop()
